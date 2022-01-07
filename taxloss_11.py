# using the anaconda path (cmd+shift+p to change interpreter)
# run this file with the play button in VSCode
# using pip install of openpyxl library

# the only change here is the addition of the KWS variable

from types import MethodDescriptorType

def realstock(row):
    if None in row[4:9] : # row 6 is prices, row 7 is returns, row 9 is weights. Check if have values.
        #print("Found None in month:", row[1].strftime(MONTHCONV))
        return 0
    return 1

def begport(msto): # chooses 100 initial stocks from given list
    minmon = msto[min(list(msto.keys()))] # list of starting-month stocks
    initst = random.sample(minmon, PORTSIZE)
    return initst

def monthlystocks(ws): # goes through entire spreadsheet to list which stocks are available for each month
    msto = {}
    r = 1 # so don't have to deal with column header
    for row in ws.values:
        if r == 1:
            None
        elif realstock(row):
            month = row[1].strftime(MONTHCONV)
            permno = row[0]
            if month in msto: # have to check because you can't append to nonexistent list
                msto[month].append(permno)
            else: 
                msto[month] = [permno]
        else:
            None
        r = r + 1
    return msto

def getret(ws, mo, stlist):
    rtdict = {}
    r = 1 # so don't have to deal with column header
    for row in ws.values:
        if r == 1:
            None
        else:
            for permno in stlist:
                pdate = row[1].strftime(MONTHCONV)
                if pdate == mo and realstock(row) :
                    if row[0] == permno:
                        ret = row[6]
                        if ret < -1:
                            ret = -1
                        rtdict[permno] = ret # 6 is returns
        r = r+1
    return rtdict

def getweights(ws, mo, beglist): # calculate weights of given permnos using market caps
    wdict = {}
    r = 1 # so don't have to deal with column header
    for row in ws.values:
        if r == 1:
            None
        else:
            pdate = row[1].strftime(MONTHCONV)
            if pdate == mo and realstock(row) :
                for permno in beglist:
                    if row[0] == permno:
                        wdict[permno] = row[8] # 8th column is market cap 
        r = r+1
    mcsum = sum(wdict.values())

    for permno in list(wdict.keys()):
        wdict[permno] = wdict[permno] / mcsum

    return wdict

def addmonths(date, nummo):
    newdate = datetime.strptime(date, MONTHCONV) + relativedelta(months = nummo)
    newdate = newdate.strftime(MONTHCONV) # i literally made so much work for myself bc i like this format
    return newdate

def addtoport(msto, month, stlist, kickedps):
    stnum = PORTSIZE-len(stlist)
    if stnum < 0:
        stnum = 0

    nextmonth = addmonths(month, 1)
    possibleadds = list(msto[nextmonth]) # to kill original reference so can edit w/o worry

    readd = []
    for permno in kickedps: # iterate through all permnos to look at reentry dates
        if (permno in possibleadds) and (kickedps[permno] != nextmonth): # check if a) it's been kicked before and b) if it's kicked period is over
            possibleadds.remove(permno)
        if kickedps[permno] == month : # i feel like i'm doing double duty but whatever
            readd.append(permno)

    for permno in readd:
        kickedps.pop(permno)

    kickedpsupdate = { stock : date for stock, date in kickedps.items() if date != month} # i'll use this later to start shortening kickedps
    kickedps = kickedpsupdate.copy()

    for permno in possibleadds:
        if (permno in stlist): # make sure not already in portfolio
            possibleadds.remove(permno)
    
    print("num possibleadds", len(possibleadds))

    if len(possibleadds) < stnum:
        toadd = possibleadds
    else: 
        # to make this more fancy, could make the new stocks be of similar market caps
        toadd = random.sample(possibleadds, stnum)

    return toadd

def outrettoxl(wsout, yb, yh, yhs) :
    # for each month, print out replaced stocks, stocks that replaced them, and then the total returns

    rtp = wsout.max_row + 1 # row to print to; remember 1-based indexing
    if rtp == 2:
        wsout.cell(row = 1, column = 1).value = "Year & Month"
        wsout.cell(row = 1, column = 2).value = "Returns of Base 100-stock Porfolio"
        wsout.cell(row = 1, column = 3).value = "Return of 100-stock Harvested Porfolio"
        wsout.cell(row = 1, column = 4).value = "Sum of Harvested Losses"

    for mon in list(yb.keys()):
        rtp = wsout.max_row + 1 # row to print to; remember 1-based indexing
        wsout.cell(row = rtp, column = 1).value = str(mon)
        wsout.cell(row = rtp, column = 2).value = yb[mon]
        wsout.cell(row = rtp, column = 3).value = yh[mon]
        wsout.cell(row = rtp, column = 4).value = yhs[mon]

    wbout.save(OUTPUTWB)
    return None

def baseport(ws, msto, months, wdict):

    yg = {} # yearly geometric returns for each stock
    yp = {} # yearly returns for portfolio
    mr = {} # monthly returns for each permno, dictionary of lists

    for m in range(1, len(months)): # for each month in the file SKIPPING the first, because of weight calculation

        rdict = getret(ws, months[m], wdict.keys())
        for permno in list(rdict.keys()): 
            if permno in list(mr.keys()):
                mr[permno].append(rdict[permno])
            else: 
                mr[permno] = [rdict[permno]]

        if m in range(13, len(months)+1, 12): # if m is a year-marker
            yp[months[m]] = 0
            for permno in list(mr.keys()): # calculate geometric return
                ygpro = 1
                for ret in mr[permno]:
                    ygpro = ygpro * (1 + ret) # running product
                yg[permno] = ygpro - 1
                yp[months[m]] = yp[months[m]] + (yg[permno] * wdict[permno]) # calculate the weighted portfolio return
                mr[permno] = []

            yp[months[m]] = yp[months[m]] * (1-TAX) # tax at END so doesn't compound
            wdict = getweights(ws, months[m], list(wdict.keys())) # recalculate market cap for next year
            print(months[m], "len mr:", len(mr.keys()), "ret", yp[months[m]])
        
        nextyr = addmonths(months[m], 11)
        if nextyr in msto.keys():
            for permno in list(wdict.keys()): # just double check it still exists
                if not(permno in msto[nextyr]):
                    wdict.pop(permno)
                    if permno in list(mr.keys()):
                        mr.pop(permno)
                    #print("one dies b4 next month")
        else:
            return yp

def harvestport(ws, msto, months, wdict):
    tokick = []
    kickedps = {}
    yhs = {} # yearly harvested sums
    yg = {} # yearly geometric returns for each stock
    yp = {} # yearly returns for portfolio
    mr = {} # monthly returns for each permno, dictionary of lists

    for m in range(1, len(months)): # for each month in the file SKIPPING the first, because of weight calculation

        rdict = getret(ws, months[m], wdict.keys())
        for permno in list(rdict.keys()): 
            if permno in list(mr.keys()):
                mr[permno].append(rdict[permno])
            else: 
                mr[permno] = [rdict[permno]]

        if m in range(13, len(months)+1, 12): # if m is a year-marker

            yp[months[m]] = 0
            yhs[months[m]] = 0
            kws = 0 # kicked-weights sum

            for permno in list(mr.keys()): # calculate geometric return
                ygpro = 1
                for ret in mr[permno]:
                    ygpro = ygpro * (1 + ret) # running product
                yg[permno] = ygpro - 1
                yp[months[m]] = yp[months[m]] + (yg[permno] * wdict[permno]) # calculate the weighted portfolio return

                if yg[permno] < CUTOFF and (wdict[permno] + kws) < KWS : # this is the 'indicator'
                    kickedps[permno] = addmonths(months[m], 12) # permno is key; assigned when stock is allowed back in
                    tokick.append(permno) # just note it b/c i don't want to mess with indexes
                    yhs[months[m]] = yhs[months[m]] + (wdict[permno] * yg[permno]) # t * w * yg[permno] --> sum them for each yg[permno]
                    kws = kws + wdict[permno]
                    print(kws)
                mr[permno] = []
            
            yp[months[m]] = yp[months[m]] * (1-TAX)
            yhs[months[m]] = yhs[months[m]] * TAX # t * w * yg[permno] --> sum them for each yg[permno]
            print(months[m], "len mr:", len(mr.keys()), "ret", yp[months[m]], "h ret", yhs[months[m]])

            for permno in tokick :
                if permno in list(wdict.keys()): # bc we h8 errors
                    wdict.pop(permno)
                if permno in list(mr.keys()):
                    mr.pop(permno)
            tokick = []
            
            toadd = addtoport(msto, months[m], list(wdict.keys()), kickedps)
            #print("len toadd", len(toadd))
            for permno in toadd:
                wdict[permno] = None #basically adding just the key to the dict

            wdict = getweights(ws, months[m], list(wdict.keys())) # recalculate market cap for next year

        nextyr = addmonths(months[m], 11)
        if nextyr in msto.keys():
            for permno in list(wdict.keys()): # just double check it still exists
                if not(permno in msto[nextyr]):
                    wdict.pop(permno)
                    if permno in list(mr.keys()):
                        mr.pop(permno)
                    #print("one dies b4 next month")
        else:
            return yp, yhs

def outret(ws, wsout): # output the returns for given sheet

    # initial setup
    msto = monthlystocks(ws) # create a DICTIONARY of all stocks in each month
    beglist = begport(msto) # choose initial 100 stocks for portfolio; list of PERMNOs

    months = list(msto.keys())
    months.sort()
    wdict = getweights(ws, months[0], beglist) # market caps and returns, outputs LIST
    #print("og weights", len(wdict))

    yb = baseport(ws, msto, months, wdict)
    yh, yhs = harvestport(ws, msto, months, wdict)
    outrettoxl(wsout, yb, yh, yhs)

    return None

# --------- BEGINNING OF CODE ------------

print()
print("------------- Begin Run -------------")
import os
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import random

tic = time.perf_counter()

dir = '/Users/kanwal/Desktop/Mason/fnan311/fnan311_taxloss/' # file directory
datadir = dir + 'taxloss_data8/' #iMPORTANT
datafiles = os.listdir(datadir)

# GLOBAL VARIABLES:
PORTSIZE = 100
TAX = .1
CUTOFF = 0
KWS = .5 # how much of the portfolio is allowed to be kicked out at one time
NUMSIM = 2
OUTPUTWB = dir + "taxloss_out15.xlsx" #output workbook name
MONTHCONV = "%Y %m"

# print(datafiles) # see what datafiles it's reading (sometimes there are hidden ones)

# create the file in which to save outputs:
wbout = Workbook()
wbout.save(OUTPUTWB)

print(datafiles)
# this loop goes through and adds a predefined column to the sheet
for i in range(len(datafiles))[3:4]: # just specific items
    currfile = datafiles[i]

    wbout.create_sheet(currfile)
    if 'Sheet' in wbout.sheetnames: # tryna get rid of the pesky automatic sheet
        wbout.remove(wbout['Sheet'])
    wsout = wbout.active

    # load current file
    wb = load_workbook(datadir + currfile)
    currsheet = wb.sheetnames[0] # only one sheet per excel 

    ws = wb[currsheet]

    for j in range(NUMSIM):
        print(currfile + '/' + currsheet, "SIM:", j) # print what file + sheet we're on
        outret(ws, wsout)

# ---end work on current file---
wbout.save(OUTPUTWB)
toc = time.perf_counter()
print(f"-------Execution time: {toc-tic:.2f}s-------")
print()
